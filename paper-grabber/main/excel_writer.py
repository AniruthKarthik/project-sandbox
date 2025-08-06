# excel_writer.py

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from typing import cast, Dict, Any
import os
from datetime import datetime
import hashlib

# Default file location - can be customized
DEFAULT_FILENAME = os.path.expanduser("~/Documents/papers/research_papers.xlsx")

def get_paper_hash(data: Dict[str, Any]) -> str:
    """Generate a hash for the paper to detect duplicates"""
    unique_string = f"{data.get('title', '')}{data.get('url', '')}"
    return hashlib.md5(unique_string.encode()).hexdigest()[:8]

def setup_worksheet_formatting(ws: Worksheet):
    """Apply professional formatting to the worksheet"""
    # Header formatting
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    
    # Apply header formatting to first row
    for col in range(1, 8):  # 7 columns
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    column_widths = {
        'A': 50,  # Title
        'B': 30,  # Authors
        'C': 10,  # Year
        'D': 12,  # Citations
        'E': 15,  # Domain
        'F': 60,  # URL
        'G': 20   # Date Added
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

def check_duplicate(ws: Worksheet, paper_hash: str) -> bool:
    """Check if paper already exists in the sheet"""
    if ws.max_row <= 1:
        return False
    
    # Check hash column (assuming it exists)
    for row in range(2, ws.max_row + 1):
        existing_hash = ws.cell(row=row, column=8).value  # Hash column
        if existing_hash == paper_hash:
            return True
    return False

def save_to_excel(data: Dict[str, Any], filename: str = None):
    """Save paper data to Excel with enhanced features"""
    if filename is None:
        filename = DEFAULT_FILENAME
    
    try:
        # Ensure the directory exists
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        
        # Generate hash for duplicate detection
        paper_hash = get_paper_hash(data)
        
        # Load or create workbook
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = cast(Worksheet, wb.active)
        else:
            wb = Workbook()
            ws = cast(Worksheet, wb.active)
            ws.title = "Research Papers"
            
            # Add headers for new workbook
            headers = ["Title", "Authors", "Year", "Citations", "Domain", "URL", "Date Added", "Hash"]
            ws.append(headers)
            setup_worksheet_formatting(ws)
        
        # Check for duplicates
        if check_duplicate(ws, paper_hash):
            print(f"Paper already exists in database: {data.get('title', 'Unknown')}")
            return False
        
        # Prepare data for insertion
        authors_str = ", ".join(data.get("authors", ["Unknown"]))
        if len(authors_str) > 250:  # Limit author string length
            authors_str = authors_str[:250] + "..."
        
        current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Add the new data row
        new_row = [
            data.get("title", "Unknown"),
            authors_str,
            data.get("year", "Unknown"),
            data.get("citations", "Unknown"),
            data.get("domain", "Unknown"),
            data.get("url", "Unknown"),
            current_date,
            paper_hash
        ]
        
        ws.append(new_row)
        
        # Apply basic formatting to new row
        row_num = ws.max_row
        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
            
            # Wrap text for title and authors
            if col in [1, 2]:  # Title and Authors columns
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Hide hash column
        ws.column_dimensions['H'].hidden = True
        
        # Save the workbook
        wb.save(filename)
        print(f"✓ Paper saved successfully to: {filename}")
        print(f"  Title: {data.get('title', 'Unknown')}")
        print(f"  Authors: {authors_str[:100]}{'...' if len(authors_str) > 100 else ''}")
        return True
        
    except PermissionError:
        print(f"Error: Cannot write to {filename}. File might be open in Excel.")
        return False
    except Exception as e:
        print(f"Error saving to Excel: {e}")
        return False

def get_paper_count(filename: str = None) -> int:
    """Get the current number of papers in the database"""
    if filename is None:
        filename = DEFAULT_FILENAME
    
    try:
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = cast(Worksheet, wb.active)
            return max(0, ws.max_row - 1)  # Subtract header row
        return 0
    except Exception:
        return 0

def export_summary(filename: str = None):
    """Print a summary of the paper database"""
    if filename is None:
        filename = DEFAULT_FILENAME
    
    try:
        if not os.path.exists(filename):
            print("No paper database found.")
            return
        
        wb = load_workbook(filename)
        ws = cast(Worksheet, wb.active)
        
        total_papers = max(0, ws.max_row - 1)
        print(f"\n📊 Paper Database Summary")
        print(f"File: {filename}")
        print(f"Total papers: {total_papers}")
        
        if total_papers > 0:
            # Show recent papers
            print(f"\nRecent additions:")
            start_row = max(2, ws.max_row - 4)  # Show last 5 papers
            for row in range(start_row, ws.max_row + 1):
                title = ws.cell(row=row, column=1).value or "Unknown"
                date_added = ws.cell(row=row, column=7).value or "Unknown"
                print(f"  • {title[:60]}{'...' if len(title) > 60 else ''} ({date_added})")
    
    except Exception as e:
        print(f"Error reading database: {e}")
