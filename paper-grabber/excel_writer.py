import os
import openpyxl
from datetime import datetime

# Define the Excel file path
EXCEL_PATH = os.path.expanduser("~/Documents/papers/research_papers.xlsx")

def ensure_directory_exists():
    """Ensure the papers directory exists"""
    papers_dir = os.path.dirname(EXCEL_PATH)
    if not os.path.exists(papers_dir):
        os.makedirs(papers_dir, exist_ok=True)
        print(f"📁 Created directory: {papers_dir}")

def save_to_excel(data):
    """Save paper data to Excel file with duplicate checking"""
    try:
        ensure_directory_exists()
        
        # Create or load workbook
        if not os.path.exists(EXCEL_PATH):
            print("📊 Creating new Excel file...")
            wb = openpyxl.Workbook()
            ws = wb.active
            # Header row matching README format
            headers = [
                "Title", "Authors", "Abstract", "Year", "Citations", 
                "DOI", "Publication", "Volume", "Issue", "Pages", 
                "EID", "URL", "Date Added"
            ]
            ws.append(headers)
            
            # Style the header row
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(
                    start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                )
        else:
            print("📊 Loading existing Excel file...")
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb.active

        # Check for duplicates by EID
        print("🔍 Checking for duplicates...")
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
            if row and len(row) > 10 and row[10] == data["eid"]:  # EID is column 11 (index 10)
                print("⚠️  Duplicate EID found. Skipping save.")
                return False

        # Prepare row data
        row_data = [
            data.get("title", "Unknown"),
            ", ".join(data.get("authors", ["Unknown"])),
            data.get("abstract", "N/A"),
            data.get("year", "Unknown"),
            str(data.get("citations", "0")),
            data.get("doi", "N/A"),
            data.get("publicationName", "N/A"),
            data.get("volume", "N/A"),
            data.get("issue", "N/A"),
            data.get("pages", "N/A"),
            data.get("eid", "Unknown"),
            data.get("url", ""),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ]
        
        # Add row to worksheet
        ws.append(row_data)
        
        # Auto-adjust column widths for better readability
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            # Set reasonable column width limits
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(EXCEL_PATH)
        print(f"✅ Paper saved to: {EXCEL_PATH}")
        
        # Show summary
        total_papers = ws.max_row - 1  # Subtract header row
        print(f"📈 Total papers in database: {total_papers}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error saving to Excel: {e}")
        import traceback
        traceback.print_exc()
        return False

def list_papers():
    """List all papers in the Excel file"""
    if not os.path.exists(EXCEL_PATH):
        print("📊 No Excel file found. No papers saved yet.")
        return
    
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        
        print(f"📚 Papers in {EXCEL_PATH}:")
        print("=" * 80)
        
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
            if row and row[0]:  # Check if title exists
                title = row[0][:60] + "..." if len(str(row[0])) > 60 else row[0]
                year = row[3] if row[3] else "Unknown"
                citations = row[4] if row[4] else "0"
                print(f"{i:3d}. {title} ({year}) - {citations} citations")
                
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")

# Test function
if __name__ == "__main__":
    # Test data
    test_data = {
        "title": "Test Paper Title",
        "authors": ["Author One", "Author Two"],
        "abstract": "This is a test abstract.",
        "year": "2023",
        "citations": "5",
        "doi": "10.1000/test",
        "publicationName": "Test Journal",
        "volume": "1",
        "issue": "2",
        "pages": "1-10",
        "eid": "2-s2.0-12345678",
        "url": "https://test.url"
    }
    
    print("Testing Excel writer...")
    save_to_excel(test_data)
    list_papers()
